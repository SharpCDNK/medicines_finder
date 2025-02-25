import os
import re
import pandas as pd
from concurrent.futures import ProcessPoolExecutor
from functools import partial
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import plotly.graph_objects as go

def process_file(file_name, input_folder, output_folder, graphs_dir, hide_quantity_columns=False):
    file_path = os.path.join(input_folder, file_name)
    print(f"Обрабатывается файл: {file_path}")

    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"Ошибка при загрузке файла {file_path}: {e}")
        return

    if 'diff_' in file_name:
        prefix = file_name.split('diff_')[-1].replace('.xlsx', '').replace('.xls', '')
    else:
        prefix = file_name.replace('.xlsx', '').replace('.xls', '')

    def extract_competitor_name(prefix):
        competitor_name_parts = []
        for part in prefix.split('_'):
            if not re.search(r'[А-Яа-я]', part):
                competitor_name_parts.append(part)
            else:
                break
        return '_'.join(competitor_name_parts)

    competitor_name = extract_competitor_name(prefix)
    quantity_columns = [col for col in df.columns if col.startswith("Количество")]

    if len(quantity_columns) < 2:
        print(f"Ошибка: недостаточно столбцов с количеством в файле {file_path}")
        return

    df[quantity_columns] = df[quantity_columns].apply(pd.to_numeric, errors='coerce').fillna(0)

    price_column = "Медианная цена"
    if price_column not in df.columns:
        print(f"Ошибка: столбец '{price_column}' отсутствует в файле {file_path}")
        return

    df[price_column] = df[price_column].astype(str).str.replace(',', '.')
    df[price_column] = df[price_column].str.extract(r'(\d+\.?\d*)', expand=False)
    df[price_column] = pd.to_numeric(df[price_column], errors='coerce').fillna(0)

    def calculate_metrics(row):
        sold = 0
        segments = []
        temp_segment = []
        prev_val = None
        segment_change_indices = []
        negative_changes = 0

        for idx, col in enumerate(quantity_columns):
            curr_val = row[col]
            if pd.isnull(curr_val):
                curr_val = 0

            if prev_val is not None:
                change = curr_val - prev_val
                if change < 0:
                    sold += -change
                    negative_changes += 1

                if curr_val > prev_val:
                    # Увеличение обнаружено, начинаем новый сегмент
                    if temp_segment and any(temp_segment):
                        segments.append(temp_segment)
                    segment_change_indices.append(idx)
                    temp_segment = []
                temp_segment.append(curr_val)
            else:
                temp_segment.append(curr_val)

            prev_val = curr_val

        if temp_segment and any(temp_segment):
            segments.append(temp_segment)

        return pd.Series({
            "Индекс изменений": sold,
            "Сегменты": len(segments),
            "Частота изменений в минус": negative_changes,
            "Индексы смены сегментов": segment_change_indices
        })

    metrics_results = df.apply(calculate_metrics, axis=1)
    df["Индекс изменений"] = metrics_results["Индекс изменений"]
    df["Сегменты"] = metrics_results["Сегменты"]
    df["Частота изменений в минус"] = metrics_results["Частота изменений в минус"]
    segment_change_indices_list = metrics_results["Индексы смены сегментов"]
    df["Заработали"] = df["Индекс изменений"] * df[price_column]

    df_filtered = df[df["Индекс изменений"] > 0].sort_values(by="Заработали", ascending=False)

    columns_to_remove = [
        'Коэффициент стабильности спроса',
        'Расчёт коэффициента стабильности',
        'КПС коэффициент плавного спроса',
        'Расчёт КПС'
    ]
    df_filtered = df_filtered.drop(columns=[col for col in columns_to_remove if col in df_filtered.columns])

    df_filtered["Ссылка на график"] = ""
    df_filtered["Название HTML"] = ""

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for index, row in df_filtered.iterrows():
        quantity_values = row[quantity_columns].values
        segment_change_indices = segment_change_indices_list.loc[row.name]

        # Строим отформатированные даты
        formatted_dates = []
        for col in quantity_columns:
            match = re.search(r'Количество.*_(\d{4})-(\d{2})-(\d{2})_(\d{2})-(\d{2})(?:.*?)$', col)
            if match:
                month = match.group(2)
                day = match.group(3)
                hour = match.group(4)
                minute = match.group(5)
                formatted_date = f"{month}-{day}_{hour}-{minute}"
                formatted_dates.append(formatted_date)
            else:
                formatted_dates.append('')

        plot_df = pd.DataFrame({
            'Дата и время': formatted_dates,
            'Количество': quantity_values
        })

        plot_df = plot_df[plot_df['Дата и время'] != '']
        plot_df['Дата и время_DT'] = pd.to_datetime(plot_df['Дата и время'], format="%m-%d_%H-%M")
        plot_df.sort_values('Дата и время_DT', inplace=True)

        key_columns = df_filtered.columns[:6]
        key_values = [str(row[col]) for col in key_columns]
        key_string = ' - '.join(key_values)

        key_string_safe = re.sub(r'\s+', '_', key_string)
        key_string_safe = re.sub(r'[<>:"/\\|?*]', '_', key_string_safe)
        max_length = 100
        if len(key_string_safe) > max_length:
            key_string_safe = key_string_safe[:max_length] + '...'

        graph_file_name = f"{prefix}_{key_string_safe}_row_{index}_graph.html"
        graph_file_path = os.path.join(graphs_dir, graph_file_name)

        fig = go.Figure()

        fig.add_annotation(
            x=0.5, y=1.15, xref='paper', yref='paper',
            showarrow=False, text=f"Конкурент: {competitor_name}",
            font=dict(size=16)
        )

        fig.add_trace(go.Scatter(
            x=plot_df['Дата и время'],
            y=plot_df['Количество'],
            mode='lines+markers',
            name='Количество',
            line_shape='spline',
            line=dict(smoothing=1.3),
            marker=dict(size=6)
        ))

        # Подсвечиваем индексы смены сегментов
        for idx in segment_change_indices:
            if idx < len(plot_df):
                fig.add_trace(go.Scatter(
                    x=[plot_df['Дата и время'].iloc[idx]],
                    y=[plot_df['Количество'].iloc[idx]],
                    mode='markers',
                    marker=dict(size=10, color='yellow'),
                    name='Новый сегмент'
                ))

        fig.update_layout(
            title=f"{key_string_safe}",
            xaxis_title="Дата и время",
            yaxis_title="Количество",
            xaxis=dict(tickangle=45),
            template='plotly_white',
            width=800,
            height=500
        )

        fig.write_html(graph_file_path)

        relative_graph_path = os.path.join("graphs", graph_file_name)
        link = '=HYPERLINK("{}", "{}")'.format(relative_graph_path, "Показать")
        df_filtered.at[index, "Ссылка на график"] = link
        df_filtered.at[index, "Название HTML"] = graph_file_name

    output_file_path = os.path.join(output_folder, f"sorted_{file_name}")

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        df_filtered.to_excel(writer, index=False, sheet_name='Data')
        workbook = writer.book
        worksheet = writer.sheets['Data']

        quantity_col_indices = [df_filtered.columns.get_loc(col) + 1 for col in df_filtered.columns if col in quantity_columns]

        for row_idx, row in enumerate(df_filtered.itertuples(), start=2):
            segment_change_indices = segment_change_indices_list.loc[row.Index]
            for idx in segment_change_indices:
                if idx < len(quantity_col_indices):
                    col_idx = quantity_col_indices[idx]
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill

        if hide_quantity_columns:
            for col_idx in quantity_col_indices:
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[col_letter].hidden = True

    print(f"Результат сохранён: {output_file_path}")

# Главный скрипт
input_folder = "Datasets/pre_ans"
output_folder = "Datasets/pre_ans_sorted_graph"
graphs_dir = os.path.join(output_folder, "graphs")

os.makedirs(output_folder, exist_ok=True)
os.makedirs(graphs_dir, exist_ok=True)

file_list = [file_name for file_name in os.listdir(input_folder) if file_name.endswith((".xls", ".xlsx"))]

with ProcessPoolExecutor() as executor:
    func = partial(
        process_file,
        input_folder=input_folder,
        output_folder=output_folder,
        graphs_dir=graphs_dir,
        hide_quantity_columns=True
    )
    executor.map(func, file_list)

print(f"Все файлы обработаны. Результаты сохранены в папке: {output_folder}")
