import pandas as pd
import os
from concurrent.futures import ProcessPoolExecutor
from functools import partial
from openpyxl.styles import PatternFill
from openpyxl import Workbook

def process_file(file_name, input_folder, output_folder):
    file_path = os.path.join(input_folder, file_name)
    print(f"Обрабатывается файл: {file_path}")
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"Ошибка при загрузке файла {file_path}: {e}")
        return

    # Поиск столбцов, начинающихся с "Количество"
    quantity_columns = [col for col in df.columns if col.startswith("Количество")]
    if len(quantity_columns) < 2:
        print(f"Ошибка: недостаточно столбцов с количеством в файле {file_path}")
        return

    # Преобразование значений столбцов количества в числовой формат и заполнение NaN нулями
    df[quantity_columns] = df[quantity_columns].apply(pd.to_numeric, errors='coerce').fillna(0)

    # Проверка наличия столбца "Медианная цена"
    price_column = "Медианная цена"
    if price_column not in df.columns:
        print(f"Ошибка: столбец '{price_column}' отсутствует в файле {file_path}")
        return

    # Преобразование столбца цены в числовой формат
    df[price_column] = df[price_column].astype(str).str.replace(',', '.')
    df[price_column] = df[price_column].str.extract(r'(\d+\.?\d*)', expand=False)
    df[price_column] = pd.to_numeric(df[price_column], errors='coerce').fillna(0)

    # Функция для расчёта "Индекс изменений", "Коэффициент стабильности спроса" и определения сегментов
    def calculate_metrics(row):
        sold = 0
        segments = []
        temp_segment = []
        prev_val = None
        segment_numbers = []
        demand_changes = []

        for idx, col in enumerate(quantity_columns):
            curr_val = row[col]
            if pd.isnull(curr_val):
                curr_val = 0

            # Расчёт изменившегося количества
            if prev_val is not None:
                change = curr_val - prev_val
                demand_changes.append(change)
                if change < 0:
                    sold += -change  # Увеличиваем проданное на уменьшение количества

                # Определение сегментов
                if change > 0:
                    if temp_segment:
                        segments.append(temp_segment)
                    temp_segment = []

            else:
                demand_changes.append(0)  # Для первого значения изменений нет

            temp_segment.append(curr_val)
            prev_val = curr_val
            segment_numbers.append(len(segments) + 1)

        if temp_segment:
            segments.append(temp_segment)

        # Расчёт коэффициента стабильности спроса
        if len(demand_changes) > 1:
            variance = pd.Series(demand_changes[1:]).var()  # Исключаем первое изменение (0)
            if variance != 0:
                stability_coefficient = 1 / variance
            else:
                stability_coefficient = float('inf')  # Если дисперсия 0, стабильность максимальна
        else:
            stability_coefficient = 0  # Недостаточно данных для расчёта

        return pd.Series({
            "Индекс изменений": sold,
            "Сегменты": len(segments),
            "Коэффициент стабильности спроса": stability_coefficient,
            "Номера сегментов": segment_numbers
        })

    # Применение функции к DataFrame
    metrics_results = df.apply(calculate_metrics, axis=1)
    df["Индекс изменений"] = metrics_results["Индекс изменений"]
    df["Сегменты"] = metrics_results["Сегменты"]
    df["Коэффициент стабильности спроса"] = metrics_results["Коэффициент стабильности спроса"]
    segment_numbers_list = metrics_results["Номера сегментов"].tolist()

    # Добавление колонки с расчётом коэффициента стабильности спроса
    df["Расчёт коэффициента стабильности"] = df["Коэффициент стабильности спроса"].apply(
        lambda x: f"1 / variance" if x != float('inf') else "1 / 0"
    )

    # Расчёт КПС (Коэффициент Плавного Спроса)
    df["КПС коэффициент плавного спроса"] = df["Индекс изменений"] / df["Сегменты"]
    df["КПС коэффициент плавного спроса"] = df["КПС коэффициент плавного спроса"].fillna(0)

    # Добавление колонки с расчётом КПС
    df["Расчёт КПС"] = df["Индекс изменений"].astype(str) + " / " + df["Сегменты"].astype(str)

    # Расчёт заработка
    df["Заработали"] = df["Индекс изменений"] * df[price_column]

    # Фильтрация и сортировка данных
    df_filtered = df[df["Индекс изменений"] > 0].sort_values(by="Заработали", ascending=False)

    # Сохранение результата с выделением ячеек
    output_file_path = os.path.join(output_folder, f"sorted_{file_name}")

    # Создаём объект ExcelWriter с использованием openpyxl
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        df_filtered.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Определяем заливку жёлтым цветом
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # Получение индексов столбцов для количества
        quantity_col_indices = [df_filtered.columns.get_loc(col) + 1 for col in quantity_columns]

        # Применение заливки к ячейкам, где сегмент меняется
        for row_idx, row in enumerate(df_filtered.itertuples(), start=2):  # Начинаем с 2, чтобы пропустить заголовок
            segment_numbers = segment_numbers_list[row.Index]
            for idx, col_idx in enumerate(quantity_col_indices):
                if idx > 0:
                    if segment_numbers[idx] != segment_numbers[idx - 1]:
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = yellow_fill
                else:
                    # Выделяем первую ячейку в строке
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = yellow_fill

    print(f"Результат сохранён в файл: {output_file_path}")

input_folder = "Datasets/pre_ans"
output_folder = "Datasets/pre_ans_sorted_new"
os.makedirs(output_folder, exist_ok=True)

file_list = [file_name for file_name in os.listdir(input_folder) if file_name.endswith((".xls", ".xlsx"))]

if __name__ == "__main__":
    with ProcessPoolExecutor() as executor:
        func = partial(process_file, input_folder=input_folder, output_folder=output_folder)
        executor.map(func, file_list)
