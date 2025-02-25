import os
import pandas as pd
import re
import multiprocessing
from collections import defaultdict
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl import Workbook

def process_competitor(competitor, analysis_dir, diff_comp_dir, output_dir, enable_correction=True):
    competitor_analysis_path = os.path.join(analysis_dir, competitor, f"{competitor}_list_for_analis.xlsx")
    competitor_diff_path = os.path.join(diff_comp_dir, competitor)

    if not os.path.exists(competitor_analysis_path) or not os.path.isdir(competitor_diff_path):
        print(f"Пропускаю {competitor}: отсутствуют необходимые файлы или папки.")
        return

    try:
        product_df = pd.read_excel(competitor_analysis_path, dtype=str).fillna('')
    except Exception as e:
        print(f"Ошибка при чтении файла {competitor_analysis_path}: {e}")
        return

    key_columns = ['name', 'item_type', 'item_form', 'prescription', 'manufacturer', 'country']
    product_df['key'] = product_df[key_columns].agg('|'.join, axis=1)
    pre_ans = product_df[['key'] + key_columns].copy()
    pre_ans['Цена min'] = 0.0
    pre_ans['Цена max'] = 0.0
    pre_ans['Медианная цена'] = 0.0

    prices_dict = defaultdict(list)

    diff_files = []
    for diff_file in os.listdir(competitor_diff_path):
        if not diff_file.endswith(('.xls', '.xlsx')):
            continue
        match = re.search(r"diff_(\d+)_parsed_data_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2})", diff_file)
        if match:
            index = int(match.group(1))
            date_time = match.group(2)
            diff_files.append((index, date_time, diff_file))

    diff_files.sort(key=lambda x: x[0])

    for index, date_time, diff_file in diff_files:
        diff_file_path = os.path.join(competitor_diff_path, diff_file)

        try:
            df = pd.read_excel(diff_file_path, dtype=str).fillna('')
        except Exception as e:
            print(f"Ошибка при чтении файла {diff_file_path}: {e}")
            continue

        if 'key' not in df.columns:
            df['key'] = df.iloc[:, :6].agg('|'.join, axis=1)

        price_column_name = 'price'
        quantity_column_name = 'only_quantity'

        if price_column_name not in df.columns or quantity_column_name not in df.columns:
            print(f"В файле {diff_file} отсутствуют колонки '{price_column_name}' или '{quantity_column_name}'.")
            continue

        quantity_info = df.set_index('key')[quantity_column_name].to_dict()

        for idx, row in df.iterrows():
            key = row['key']
            price_str = str(row[price_column_name])

            match_price = re.search(r"([\d\s]+[.,]?\d*)", price_str)
            if match_price:
                price_value = match_price.group(1)
                price_value = price_value.replace(',', '.').replace(' ', '')
                try:
                    price = float(price_value)
                    prices_dict[key].append(price)
                except (ValueError, TypeError):
                    continue

        # Преобразуем количества в float
        quantity_info = {k: float(v) if v != '' else 0.0 for k, v in quantity_info.items()}

        pre_ans[f"Количество {index}_{date_time}"] = pre_ans['key'].map(quantity_info).fillna(0.0)

    if not prices_dict:
        print("Внимание: словарь цен пуст. Проверьте корректность обработки цен.")

    def compute_price_statistics(prices):
        if prices:
            return min(prices), max(prices), np.median(prices)
        return 0.0, 0.0, 0.0

    stats = pre_ans['key'].apply(lambda k: compute_price_statistics(prices_dict.get(k, [])))
    pre_ans['Цена min'] = stats.apply(lambda x: x[0])
    pre_ans['Цена max'] = stats.apply(lambda x: x[1])
    pre_ans['Медианная цена'] = stats.apply(lambda x: x[2])

    corrected_cells = []  # Инициализируем список исправленных ячеек

    if enable_correction:
        # Исправляем значения и получаем список исправленных ячеек
        pre_ans, corrected_cells = check_and_correct_values(pre_ans)
    else:
        # Если исправление отключено, инициализируем пустой список исправленных ячеек
        corrected_cells = []

    pre_ans.drop(columns=['key'], inplace=True)
    output_file = os.path.join(output_dir, f"pre_ans_{competitor}.xlsx")

    # Сохраняем DataFrame с выделением ячеек
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        pre_ans.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Применяем серый цвет к ячейкам, которые нужно выделить
        gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        for cell_pos in corrected_cells:
            row_idx, col_idx = cell_pos
            cell = worksheet.cell(row=row_idx + 2, column=col_idx + 1)  # +2 из-за заголовка и 0-й индексации
            cell.fill = gray_fill

    print(f"\npre_ans сохранен: {output_file}")

def check_and_correct_values(df):
    quantity_columns = [col for col in df.columns if col.startswith('Количество')]
    corrected_cells = []  # Список для хранения координат исправленных ячеек

    num_columns = len(quantity_columns)
    for idx, row in df.iterrows():
        values = [float(row[col]) for col in quantity_columns]
        i = 0
        while i < num_columns:
            p1 = values[i]
            if p1 > 0:
                start_idx = i
                max_intermediate_positions = min(8, num_columns - i - 2)
                found_p2 = False
                for offset in range(1, max_intermediate_positions + 2):
                    j = i + offset
                    if j >= num_columns:
                        break
                    p2 = values[j]
                    if p2 > 0:
                        if p2 <= p1:
                            # Заменяем промежуточные значения на p2
                            for k in range(start_idx + 1, j):
                                values[k] = p2
                                col_name = quantity_columns[k]
                                df.at[idx, col_name] = p2
                                corrected_cells.append((idx, df.columns.get_loc(col_name)))
                            i = j  # Продолжаем поиск с позиции p2
                            found_p2 = True
                            break
                        else:
                            # Если p2 > p1, дальше искать нет смысла
                            i = j
                            found_p2 = True
                            break
                if not found_p2:
                    # Если подходящая p2 не найдена, переходим к следующему значению
                    i += 1
            else:
                i += 1

    return df, corrected_cells

def generate_pre_ans_table(analysis_dir, diff_comp_dir, output_dir, enable_correction=True):
    if not os.path.exists(analysis_dir):
        print(f"Папка {analysis_dir} не существует.")
        return
    if not os.path.exists(diff_comp_dir):
        print(f"Папка {diff_comp_dir} не существует.")
        return
    os.makedirs(output_dir, exist_ok=True)

    competitors = [comp for comp in os.listdir(analysis_dir) if os.path.isdir(os.path.join(analysis_dir, comp))]

    with multiprocessing.Pool(processes=multiprocessing.cpu_count()) as pool:
        pool.starmap(process_competitor, [(comp, analysis_dir, diff_comp_dir, output_dir, enable_correction) for comp in competitors])

analysis_directory = "Datasets/list_for_analis"
comparison_directory = "Datasets/diff_comp"
output_directory = "Datasets/pre_ans"

enable_correction = True

generate_pre_ans_table(analysis_directory, comparison_directory, output_directory, enable_correction=enable_correction)
